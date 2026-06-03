from flask import Flask, render_template, request, jsonify, send_file
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, Reference
import io
import json
import datetime

app = Flask(__name__)

MONTHS_SHORT = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
MONTHS_FULL  = ["January","February","March","April","May","June",
                "July","August","September","October","November","December"]
MONTH_NAME_MAP = {m.lower():i+1 for i,m in enumerate(MONTHS_FULL)}
MONTH_NAME_MAP.update({m.lower():i+1 for i,m in enumerate(MONTHS_SHORT)})

LY_LIFTS = {"TPR":0.18,"Display":0.24,"BOGO":0.31,"Feature Ad":0.22}

SAMPLE_LY = {
    1:{"fv":501736,"af":584286}, 2:{"fv":510904,"af":556387},
    3:{"fv":613481,"af":570698}, 4:{"fv":654845,"af":654845},
    5:{"fv":710039,"af":710039}, 6:{"fv":753851,"af":753851},
    7:{"fv":558666,"af":558666}, 8:{"fv":665828,"af":665828},
    9:{"fv":588604,"af":588604}, 10:{"fv":686701,"af":686701},
    11:{"fv":633831,"af":633831},12:{"fv":520920,"af":520920},
}

def parse_ly_file(file):
    try:
        df = pd.read_excel(file, header=None)
    except Exception:
        file.seek(0)
        df = pd.read_csv(file, header=None)

    # Find header row
    header_row = None
    for i, row in df.iterrows():
        vals = [str(v).lower().strip() for v in row]
        if any("month" in v for v in vals) and any("forecast" in v or "volume" in v or "package" in v for v in vals):
            header_row = i
            break

    if header_row is None:
        return None, "Could not find a header row with Month and Forecast columns."

    df.columns = df.iloc[header_row]
    df = df.iloc[header_row+1:].reset_index(drop=True)
    df.columns = [str(c).strip() for c in df.columns]

    # Find columns
    col_map = {}
    for col in df.columns:
        cl = col.lower()
        if "month" in cl: col_map["month"] = col
        elif "package" in cl or "pkg" in cl: col_map["package"] = col
        elif "actual" in cl and "forecast" in cl: col_map["af"] = col
        elif "forecast" in cl and "volume" in cl: col_map["fv"] = col
        elif "forecast" in cl and "fv" not in col_map: col_map["fv"] = col
        elif "volume" in cl and "fv" not in col_map: col_map["fv"] = col

    if "month" not in col_map: return None, "Could not find a Month column."
    if "fv" not in col_map: return None, "Could not find a Forecast Volume column."

    # Filter for Monster 16oz Single if package col exists
    if "package" in col_map:
        df = df[df[col_map["package"]].astype(str).str.lower().str.contains("monster") &
                df[col_map["package"]].astype(str).str.lower().str.contains("16") &
                df[col_map["package"]].astype(str).str.lower().str.contains("single")]

    if df.empty: return None, "No Monster 16oz Single rows found after filtering."

    result = {}
    for _, row in df.iterrows():
        mo_raw = str(row[col_map["month"]]).strip()
        try:
            mo = int(float(mo_raw))
        except ValueError:
            mo = MONTH_NAME_MAP.get(mo_raw.lower()[:9])
        if not mo or mo < 1 or mo > 12: continue
        try:
            fv = float(row[col_map["fv"]])
        except (ValueError, TypeError):
            continue
        af = fv
        if "af" in col_map:
            try: af = float(row[col_map["af"]])
            except (ValueError, TypeError): pass
        result[mo] = {"fv": round(fv, 2), "af": round(af, 2)}

    if not result: return None, "No valid month rows found."
    return result, None


def build_excel(config):
    retailer = config.get("retailer", "Retailer")
    brand    = "Monster 16oz Single"
    year     = int(config.get("year", 2026))
    growth   = float(config.get("growth", 5)) / 100
    preparer = config.get("preparer", "")
    ly_data  = {int(k): v for k, v in config.get("ly_data", SAMPLE_LY).items()}
    promos   = config.get("promos", [])

    mo_promo = {}
    for p in promos:
        mo = int(p["month"])
        ly_lift = LY_LIFTS.get(p["type"])
        lift = ly_lift if ly_lift is not None else float(p.get("manualLift") or 0) / 100
        mo_promo[mo] = {"type": p["type"], "lift": lift, "source": "LY" if ly_lift is not None else "Manual"}

    # Colors
    NAVY="1B2A4A"; MID="2E5090"; LT_BLUE="D6E4F0"; TEAL="0F6E56"
    LT_TEAL="E1F5EE"; AMBER="854F0B"; LT_AMBER="FAEEDA"
    WHITE="FFFFFF"; LT_GRAY="F5F7FA"; DARK="1A1A2E"

    def fl(c): return PatternFill("solid", start_color=c, fgColor=c)
    def fn(bold=False,size=10,color=DARK): return Font(name="Arial",bold=bold,size=size,color=color)
    def al(h="left",v="center",indent=0): return Alignment(horizontal=h,vertical=v,indent=indent)
    def bd():
        s=Side(style="thin",color="C0C8D8")
        return Border(left=s,right=s,top=s,bottom=s)

    def hdr(ws,r,c,val,bg=NAVY,fg=WHITE,sz=10,h="center"):
        cell=ws.cell(r,c,val); cell.font=fn(True,sz,fg)
        cell.fill=fl(bg); cell.alignment=al(h,"center"); cell.border=bd()
        return cell

    def dc(ws,r,c,val,bg=WHITE,fg=DARK,bold=False,fmt=None,h="left"):
        cell=ws.cell(r,c,val); cell.font=fn(bold,10,fg)
        cell.fill=fl(bg); cell.alignment=al(h,"center",indent=1 if h=="left" else 0)
        cell.border=bd()
        if fmt: cell.number_format=fmt
        return cell

    total_ly  = sum(v["fv"] for v in ly_data.values())
    total_ly_af = sum(v["af"] for v in ly_data.values())
    total_base = sum(round(ly_data[m]["fv"]*(1+growth)) for m in range(1,13) if m in ly_data)
    total_promo_lift = 0
    for mo, p in mo_promo.items():
        if mo in ly_data:
            base_mo = round(ly_data[mo]["fv"]*(1+growth))
            total_promo_lift += round(base_mo * p["lift"])
    total_fcst = total_base + total_promo_lift

    wb = openpyxl.Workbook()

    # ── Sheet 1: Dashboard ──────────────────────────────
    sv = wb.active; sv.title = "Dashboard"
    sv.sheet_view.showGridLines = False
    for col,w in [("A",2),("B",16),("C",16),("D",16),("E",16),("F",16),("G",16),("H",2)]:
        sv.column_dimensions[col].width = w

    sv.row_dimensions[1].height=6; sv.row_dimensions[2].height=52; sv.row_dimensions[3].height=6
    for c in range(1,9): sv.cell(2,c).fill=fl(NAVY)
    sv.merge_cells("B2:G2")
    tc=sv["B2"]; tc.value=f"MONSTER 16OZ SINGLE  —  {year} FORECAST"
    tc.font=Font(name="Arial",bold=True,size=17,color=WHITE); tc.alignment=al("left","center",indent=1)

    sv.row_dimensions[4].height=22
    for c in range(1,9): sv.cell(4,c).fill=fl(MID)
    sv.merge_cells("B4:D4")
    sv["B4"].value=f"{retailer}  |  {preparer}"
    sv["B4"].font=fn(True,9,WHITE); sv["B4"].alignment=al("left","center",indent=1)
    sv.merge_cells("E4:G4")
    sv["E4"].value=f"Growth: {growth:.1%}   |   {datetime.date.today().strftime('%b %d, %Y')}"
    sv["E4"].font=fn(False,9,WHITE); sv["E4"].alignment=al("right","center")

    sv.row_dimensions[5].height=12

    kpis=[("LY Fcst Volume",total_ly,MID,LT_BLUE),
          ("LY Act + Fcst",total_ly_af,MID,LT_BLUE),
          ("Base Forecast",total_base,TEAL,LT_TEAL),
          ("Promo Lift $",total_promo_lift,TEAL,LT_TEAL),
          ("TY Forecast",total_fcst,NAVY,LT_BLUE),
          ("$ vs LY",total_fcst-total_ly,NAVY,LT_BLUE)]
    sv.row_dimensions[6].height=18; sv.row_dimensions[7].height=34; sv.row_dimensions[8].height=12

    for i,(label,val,dark_c,light_c) in enumerate(kpis):
        col=i+2
        lc=sv.cell(6,col,label); lc.font=fn(True,9,dark_c); lc.fill=fl(light_c)
        lc.alignment=al("center","center"); lc.border=bd()
        vc=sv.cell(7,col,val); vc.font=fn(True,13,dark_c); vc.fill=fl(light_c)
        vc.alignment=al("center","center"); vc.number_format="$#,##0"; vc.border=bd()

    sv.row_dimensions[9].height=8; sv.row_dimensions[10].height=22
    for ci,h in enumerate(["","Month","LY Fcst Vol","Base Fcst","Promo","TY Forecast","% vs LY",""],1):
        if h:
            c=sv.cell(10,ci,h); c.font=fn(True,9,WHITE); c.fill=fl(NAVY)
            c.alignment=al("center","center"); c.border=bd()

    for mi in range(1,13):
        r=10+mi; sv.row_dimensions[r].height=20
        bg=WHITE if mi%2==0 else LT_GRAY
        ly_fv=ly_data.get(mi,{}).get("fv",0)
        base=round(ly_fv*(1+growth))
        promo=mo_promo.get(mi)
        lift_amt=round(base*promo["lift"]) if promo else 0
        final=base+lift_amt
        pct=((final/ly_fv)-1)*100 if ly_fv else 0

        dc(sv,r,2,MONTHS_SHORT[mi-1],bg,DARK,True,h="center")
        dc(sv,r,3,ly_fv,bg,DARK,fmt="$#,##0",h="right")
        dc(sv,r,4,base,bg,DARK,fmt="$#,##0",h="right")
        if promo:
            pt=sv.cell(r,5,f"{promo['type']} +{round(promo['lift']*100)}%")
            pt.font=Font(name="Arial",bold=True,size=9,color=TEAL)
            pt.fill=fl(LT_TEAL); pt.alignment=al("center","center"); pt.border=bd()
        else:
            cell=sv.cell(r,5,"—"); cell.font=fn(False,9,"AAAAAA")
            cell.fill=fl(bg); cell.alignment=al("center","center"); cell.border=bd()
        fc=sv.cell(r,6,final); fc.font=fn(True,9,TEAL if promo else NAVY)
        fc.fill=fl(LT_TEAL if promo else bg); fc.alignment=al("right","center")
        fc.number_format="$#,##0"; fc.border=bd()
        pc=sv.cell(r,7,pct/100); pc.font=fn(False,9,"0F6E56" if pct>=0 else "A32D2D")
        pc.fill=fl(bg); pc.alignment=al("center","center")
        pc.number_format="+0.0%;-0.0%"; pc.border=bd()

    tot_r=23; sv.row_dimensions[tot_r].height=24
    for ci in range(1,9): sv.cell(tot_r,ci).fill=fl(NAVY); sv.cell(tot_r,ci).border=bd()
    sv.cell(tot_r,2).value="TOTAL"; sv.cell(tot_r,2).font=fn(True,10,WHITE); sv.cell(tot_r,2).alignment=al("center","center")
    for col,val in [(3,total_ly),(4,total_base),(6,total_fcst),(7,(total_fcst/total_ly-1) if total_ly else 0)]:
        c=sv.cell(tot_r,col,val); c.font=fn(True,10,WHITE); c.fill=fl(NAVY)
        c.alignment=al("right","center")
        c.number_format="$#,##0" if col!=7 else "+0.0%;-0.0%"; c.border=bd()

    # ── Sheet 2: Monthly Detail ─────────────────────────
    md=wb.create_sheet("Monthly Detail"); md.sheet_view.showGridLines=False
    for col,w in [("A",2),("B",16),("C",16),("D",16),("E",16),("F",14),("G",14),("H",14),("I",2)]:
        md.column_dimensions[col].width=w
    md.row_dimensions[1].height=6; md.row_dimensions[2].height=46
    for c in range(1,10): md.cell(2,c).fill=fl(NAVY)
    md.merge_cells("B2:H2")
    tc=md["B2"]; tc.value="MONTHLY DETAIL"; tc.font=Font(name="Arial",bold=True,size=15,color=WHITE)
    tc.alignment=al("left","center",indent=1)
    md.row_dimensions[3].height=8; md.row_dimensions[4].height=22
    for ci,h in enumerate(["","Month","LY Fcst Vol","LY Act+Fcst","Base Fcst","Promo Type","Lift %","TY Forecast",""],1):
        if h:
            c=md.cell(4,ci,h); c.font=fn(True,9,WHITE); c.fill=fl(NAVY)
            c.alignment=al("center","center"); c.border=bd()
    for mi in range(1,13):
        r=4+mi; md.row_dimensions[r].height=20
        bg=WHITE if mi%2==0 else LT_GRAY
        ly_fv=ly_data.get(mi,{}).get("fv",0); ly_af=ly_data.get(mi,{}).get("af",0)
        base=round(ly_fv*(1+growth)); promo=mo_promo.get(mi)
        lift_amt=round(base*promo["lift"]) if promo else 0; final=base+lift_amt
        dc(md,r,2,MONTHS_FULL[mi-1],bg,DARK,True,h="center")
        dc(md,r,3,ly_fv,bg,DARK,fmt="$#,##0",h="right")
        dc(md,r,4,ly_af,bg,DARK,fmt="$#,##0",h="right")
        dc(md,r,5,base,bg,DARK,fmt="$#,##0",h="right")
        if promo:
            pt=md.cell(r,6,promo["type"]); pt.font=Font(name="Arial",bold=True,size=9,color=TEAL)
            pt.fill=fl(LT_TEAL); pt.alignment=al("center","center"); pt.border=bd()
            pl=md.cell(r,7,promo["lift"]); pl.font=Font(name="Arial",bold=True,size=9,color=TEAL)
            pl.fill=fl(LT_TEAL); pl.alignment=al("center","center"); pl.number_format="0%"; pl.border=bd()
        else:
            for c in [6,7]:
                cell=md.cell(r,c,"—"); cell.font=fn(False,9,"AAAAAA")
                cell.fill=fl(bg); cell.alignment=al("center","center"); cell.border=bd()
        fc=md.cell(r,8,final); fc.font=fn(True,9,TEAL if promo else NAVY)
        fc.fill=fl(LT_TEAL if promo else LT_BLUE); fc.alignment=al("right","center")
        fc.number_format="$#,##0"; fc.border=bd()
    tot_r=17; md.row_dimensions[tot_r].height=24
    for ci in range(1,10): md.cell(tot_r,ci).fill=fl(NAVY); md.cell(tot_r,ci).border=bd()
    md.cell(tot_r,2).value="TOTAL"; md.cell(tot_r,2).font=fn(True,10,WHITE); md.cell(tot_r,2).alignment=al("center","center")
    for col,val in [(3,total_ly),(4,total_ly_af),(5,total_base),(8,total_fcst)]:
        c=md.cell(tot_r,col,val); c.font=fn(True,10,WHITE); c.fill=fl(NAVY)
        c.alignment=al("right","center"); c.number_format="$#,##0"; c.border=bd()

    # ── Sheet 3: Promo Summary ──────────────────────────
    if promos:
        ps=wb.create_sheet("Promo Summary"); ps.sheet_view.showGridLines=False
        for col,w in [("A",2),("B",18),("C",14),("D",12),("E",14),("F",16),("G",16),("H",2)]:
            ps.column_dimensions[col].width=w
        ps.row_dimensions[1].height=6; ps.row_dimensions[2].height=44
        for c in range(1,9): ps.cell(2,c).fill=fl(NAVY)
        ps.merge_cells("B2:G2")
        tc=ps["B2"]; tc.value="PROMO LIFT SUMMARY"
        tc.font=Font(name="Arial",bold=True,size=14,color=WHITE); tc.alignment=al("left","center",indent=1)
        ps.row_dimensions[3].height=8; ps.row_dimensions[4].height=22
        for ci,h in enumerate(["","Promo Type","Month","Lift %","Lift Source","Est. Incremental $",""],1):
            if h:
                c=ps.cell(4,ci,h); c.font=fn(True,9,WHITE); c.fill=fl(NAVY)
                c.alignment=al("center","center"); c.border=bd()
        total_inc=0
        for i,p in enumerate(promos):
            r=5+i; ps.row_dimensions[r].height=22
            mo=int(p["month"]); ly_lift=LY_LIFTS.get(p["type"])
            lift=ly_lift if ly_lift is not None else float(p.get("manualLift") or 0)/100
            is_ly=ly_lift is not None
            base_mo=round(ly_data.get(mo,{}).get("fv",0)*(1+growth))
            inc=round(base_mo*lift); total_inc+=inc
            dc(ps,r,2,p["type"],LT_TEAL,TEAL,True,h="center")
            dc(ps,r,3,MONTHS_SHORT[mo-1],LT_GRAY,DARK,h="center")
            lc=ps.cell(r,4,lift); lc.font=Font(name="Arial",bold=True,size=10,color=TEAL)
            lc.fill=fl(LT_TEAL); lc.alignment=al("center","center"); lc.number_format="0%"; lc.border=bd()
            sc=ps.cell(r,5,"From LY" if is_ly else "Manual")
            sc.font=Font(name="Arial",bold=True,size=9,color=TEAL if is_ly else AMBER)
            sc.fill=fl(LT_TEAL if is_ly else LT_AMBER); sc.alignment=al("center","center"); sc.border=bd()
            ic=ps.cell(r,6,inc); ic.font=Font(name="Arial",bold=True,size=10,color=TEAL)
            ic.fill=fl(LT_TEAL); ic.alignment=al("right","center"); ic.number_format="$#,##0"; ic.border=bd()
        tot_r=5+len(promos); ps.row_dimensions[tot_r].height=24
        for ci in range(1,9): ps.cell(tot_r,ci).fill=fl(NAVY); ps.cell(tot_r,ci).border=bd()
        ps.cell(tot_r,2).value="TOTAL"; ps.cell(tot_r,2).font=fn(True,10,WHITE); ps.cell(tot_r,2).alignment=al("center","center")
        tc2=ps.cell(tot_r,6,total_inc); tc2.font=fn(True,10,WHITE); tc2.fill=fl(NAVY)
        tc2.alignment=al("right","center"); tc2.number_format="$#,##0"; tc2.border=bd()

    # Tab colors
    sv.sheet_properties.tabColor=NAVY
    md.sheet_properties.tabColor=TEAL
    wb.active=sv

    output=io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


@app.route("/")
def index():
    return render_template("index.html", ly_lifts=json.dumps(LY_LIFTS))

@app.route("/upload", methods=["POST"])
def upload():
    f = request.files.get("file")
    if not f: return jsonify({"error": "No file received"}), 400
    data, err = parse_ly_file(f)
    if err: return jsonify({"error": err}), 400
    return jsonify({"success": True, "data": data, "months": len(data)})

@app.route("/generate", methods=["POST"])
def generate():
    config = request.get_json()
    output = build_excel(config)
    retailer = config.get("retailer","Forecast").replace(" ","_")
    year = config.get("year", 2026)
    filename = f"{retailer}_Monster16oz_{year}.xlsx"
    return send_file(output, as_attachment=True,
                     download_name=filename,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

if __name__ == "__main__":
    app.run(debug=True, port=5000)
